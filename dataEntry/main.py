from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib


bg_color = "#326273"
text_color = "#fff"

file = pathlib.Path("Backend_data.xlsx")

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1']="Full Name"
    sheet['B1']="PhoneNumber"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    file.save("Backend_data.xlsx")

def addData():

   name = nameValue.get()
   contact= contactValue.get()
   age=ageValue.get()
   gender=genderEntry.get()
   address=addressEntry.get(1.0,END)

   if name !="" and contact != "" and age != "" and address != "":
       file = openpyxl.load_workbook("Backend_data.xlsx")
       sheet = file.active
       sheet.cell(column=1, row=sheet.max_row+1, value=name)
       sheet.cell(column=2, row=sheet.max_row, value=contact)
       sheet.cell(column=3, row=sheet.max_row, value=age)
       sheet.cell(column=4, row=sheet.max_row, value=gender)
       sheet.cell(column=5, row=sheet.max_row, value=address)

       file.save(r'Backend_data.xlsx')

       clear()
   else:
       messagebox.showerror("Warning", "All File Required")



def clear():
    nameValue.set("")
    contactValue.set("")
    ageValue.set("")
    genderEntry.set("Select")
    addressEntry.delete(1.0, END)







root = Tk()
root.title("Data Entry Form")
root.geometry("700x400+300+200")
root.resizable(False,False)
root.configure(bg=bg_color)


#icon 
# icon_image = PhotoImage(file="")
# root.iconphoto(False,icon_image)

#heading 
Label(root, text="Please fill out this Entry Form: ", font="arial 13", bg=bg_color, fg=text_color).place(x=20, y=50)


#Entry Fild labeles
Label(root, text="Name",font=20, bg=bg_color, fg=text_color).place(x=50, y=100)
Label(root, text="Contact No", font=20, bg=bg_color, fg=text_color).place(x=50, y=150)
Label(root,text="Age", font=20, bg=bg_color, fg=text_color).place(x=50, y=200)
Label(root, text="Gender", font=20, bg=bg_color, fg=text_color).place(x=380, y=200)
Label(root, text="Address", font=20, bg=bg_color, fg=text_color).place(x=50, y=250)

#Entry fild input 
options=["Male", "Female"]
nameValue = StringVar()
contactValue = StringVar()
ageValue=StringVar()

nameEntry=Entry(root, textvariable=nameValue, width=40, bd=2, font=20).place(x=200, y=100)
contactEntry = Entry(root, textvariable=contactValue, width=40, bd=2, font=20).place(x=200, y=150)
ageEntry = Entry(root, textvariable=ageValue, width=15, bd=2, font=15).place(x=200, y=200)
addressEntry = Text(root, width=55, height=4, bd=4)
addressEntry.place(x=200, y=250)

genderEntry = Combobox(root,values=options, width=15, state="r", font="arial 15")
genderEntry.set("Select")
genderEntry.place(x=455, y=200)


#button

Button(root, text="Submit", bg=bg_color, fg=text_color, width=15, height=2, bd=2, command=addData).place(x=210, y=340)
Button(root, text="Clear", bg=bg_color, fg=text_color, width=15, height=2, bd=2, command=clear).place(x=360, y=340)
Button(root, text="Exit", bg=bg_color, fg=text_color, width=15, height=2, bd=2, command=lambda: root.destroy()).place(x=510, y=340)





root.mainloop()
